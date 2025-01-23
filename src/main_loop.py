import requests
import csv
import os
from datetime import datetime, timedelta
from supabase import create_client, Client
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import time
import logging

load_dotenv("config/settings.env")
logging.basicConfig(
    filename=os.getenv("LOG_FILE_PATH"),
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)

THINGSPEAK_URL = f"https://api.thingspeak.com/channels/2769816/feeds.json?api_key=WTIFG3EJDHOGBRKF&results=1"
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_API_KEY = os.getenv("SUPABASE_API_KEY")
SUPABASE_TABLE_NAME = os.getenv("SUPABASE_TABLE_NAME")
CSV_FILE_PATH = os.getenv("CSV_FILE_PATH")
# EXCEL_FILE_PATH = os.getenv("EXCEL_FILE_PATH")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_API_KEY)

last_entry_id = None
last_processed_time = datetime.min

def fetch_data_from_thingspeak():
    try:
        response = requests.get(THINGSPEAK_URL, timeout=10)
        response.raise_for_status()
        data = response.json()
        time.sleep(60)
        if not data["feeds"]:
            logging.warning("No data found in ThingSpeak.")
            return None
        return data["feeds"][-1]
    except Exception as e:
        logging.error(f"Error fetching data: {e}")
        return None
    

def send_data_to_supabase(data):
    try:
        formatted_data = {
            "entry_id": data["entry_id"],
            "created_at": data["created_at"],
            "ph": float(data.get("field1", 0)),
            "do": float(data.get("field2", 0)),
            "t": float(data.get("field3", 0)),
            "v": float(data.get("field4", 0)),
            "sensor_id": data.get("field5", "unknown")
        }
        response = supabase.table(SUPABASE_TABLE_NAME).insert(formatted_data).execute()
        if response.data:
            logging.info("Data successfully sent to Supabase.")
        else:
            logging.error(f"Error sending data to Supabase: {response}")
    except Exception as e:
        logging.error(f"Supabase error: {e}")

def save_data_to_csv(data):
    file_exists = os.path.exists(CSV_FILE_PATH)
    with open(CSV_FILE_PATH, "a", newline="") as file:
        writer = csv.writer(file)
        if not file_exists:
            writer.writerow(["Entry ID", "Created At", "PH", "DO", "T", "V", "Sensor ID"])
        writer.writerow([
            data.get("entry_id"), data.get("created_at"),
            data.get("field1"), data.get("field2"),
            data.get("field3"), data.get("field4"),
            data.get("field5")
        ])
    logging.info("Data saved to CSV.")

# def save_data_to_excel(data):
#     file_exists = os.path.exists(EXCEL_FILE_PATH)
#     if file_exists:
#         workbook = load_workbook(EXCEL_FILE_PATH)
#         sheet = workbook.active
#     else:
#         workbook = Workbook()
#         sheet = workbook.active
#         sheet.title = "Sensor Data"
#         sheet.append(["Entry ID", "Created At", "PH", "DO", "T", "V", "Sensor ID"])
#         logging.info("Excel file created with headers.")

#     sheet.append([
#         data["entry_id"], data["created_at"],
#         data.get("field1"), data.get("field2"),
#         data.get("field3"), data.get("field4"),
#         data.get("field5")
#     ])

#     if not file_exists:
#         tab = Table(displayName="SensorTable", ref=f"A1:G{sheet.max_row}")
#         style = TableStyleInfo(
#             name="TableStyleMedium9",
#             showFirstColumn=False,
#             showLastColumn=False,
#             showRowStripes=True,
#             showColumnStripes=True
#         )
#         tab.tableStyleInfo = style
#         sheet.add_table(tab)

#         for cell in sheet["1:1"]:
#             cell.font = Font(bold=True)
#             cell.alignment = Alignment(horizontal="center")

#     workbook.save(EXCEL_FILE_PATH)
#     workbook.close()
#     logging.info("Data successfully saved to Excel.")

def main():
    global last_entry_id, last_processed_time
    while True:
        try:
            logging.info("Fetching data...")
            data = fetch_data_from_thingspeak()
            if data:
                current_entry_id = data["entry_id"]
                current_time = datetime.now()
                if current_entry_id != last_entry_id and (current_time - last_processed_time) >= timedelta(seconds=60):
                    send_data_to_supabase(data)
                    save_data_to_csv(data)
                    # save_data_to_excel(data)
                    last_entry_id = current_entry_id
                    last_processed_time = current_time
                    logging.info(f"New data with entry_id {current_entry_id} processed successfully.")
                else:
                    logging.info(f"No new entry or waiting for the next interval (entry_id: {current_entry_id}).")
            else:
                logging.info("No data fetched. Waiting for next iteration...")
            time.sleep(60)
        except KeyboardInterrupt:
            logging.info("Program stopped by user.")
            break
        except Exception as e:
            logging.error(f"Unexpected error: {e}")

if __name__ == "__main__":
    main()
